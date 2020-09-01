from openpyxl import load_workbook


def main():
    wb = load_workbook("./sheetf.xlsx")  # 讀取excel檔案
    # print(wb.sheetnames)  # display all sheet name
    for sheet in wb.worksheets:
        # print(sheet.sheet_properties)  #
        # print(sheet.sheet_properties.codeName, sheet.sheet_properties.tabColor)

        # 按照sheet的顏色篩選，先判斷None，預設的sheet顏色是None；tabColor回傳一個Color物件
        if sheet.sheet_properties.tabColor is None or sheet.sheet_properties.tabColor.theme != 7:
            # wb.remove(sheet)  # 移除不要的sheet
            sheet.sheet_state = "hidden"  # 隱藏不要的sheet

    wb.save("./fc.xlsx")  # 存新檔


if __name__ == "__main__":
    main()
